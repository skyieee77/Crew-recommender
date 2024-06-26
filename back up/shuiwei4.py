# -*- coding: UTF-8 -*-
# -*- coding: UTF-8 -*-
# -*- coding: UTF-8 -*-
import openpyxl
from openpyxl.styles import Font
import pandas as pd


# 打开现有的Excel文件  todo 修改文件名

# file_path = 'new.xlsx'
file_path = '../data/调试20231018(v2).xlsx'
workbook = openpyxl.load_workbook(file_path)

# 选择要写入的工作表  todo 修改sheet名
sheet = workbook['6分钟分步水量控制模型']

# 让用户指定要插入数据的位置  todo 修改位置 部分位置写不进去，可以插入到文件中空白位置再粘贴
start_row = 11
start_col = 12


target_level1 = float(94.8)
current_level1 = float(input("请输入麻峪当前水位（按Enter键结束输入）: "))

# 将变量的值赋给特定单元格
value_to_insert1 = current_level1
sheet['D12'] = value_to_insert1  # 这里将值插入到A1单元格



if (current_level1>target_level1):
    current_level1 = float(current_level1 + 0.03)
else:
    current_level1 = float(current_level1 - 0.03)


# 杏石口64.3有问题 todo
target_level2 =float(64.5)
current_level2 = float(input("请输入杏石口当前水位（按Enter键结束输入）: "))
if (current_level2>target_level2):
    current_level2 = float(current_level2 + 0.03)
else:
    current_level2 = float(current_level2 - 0.03)

value_to_insert2 = current_level2
sheet['E12'] = value_to_insert2  # 这里将值插入到A1单元格




def calculate_adjustment1(target_level1, current_level1, normal_descent=0.03):
    steps = 0
    adjusted_level = current_level1
    data1 = []

    while abs(adjusted_level - target_level1) > 0.001:
        if target_level1 > adjusted_level:
            step_adjustment = min(normal_descent, target_level1 - adjusted_level)
        else:
            step_adjustment = max(-normal_descent, target_level1 - adjusted_level)

        adjusted_level += step_adjustment
        steps += 1
        data1.append(adjusted_level)

    return data1


def calculate_adjustment2(target_level2, current_level2, normal_descent=0.03):
    steps = 0
    adjusted_level = current_level2
    data2 = []

    while abs(adjusted_level - target_level2) > 0.001:
        if target_level2 > adjusted_level:
            step_adjustment = min(normal_descent, target_level2 - adjusted_level)
        else:
            step_adjustment = max(-normal_descent, target_level2 - adjusted_level)

        adjusted_level += step_adjustment
        steps += 1
        data2.append(adjusted_level)

    return data2


# # 模块1和模块2的数据生成 todo 修改
# current_level1 = 95.6
# target_level1 = 94.8
# # current_level1 = float(input("请输入麻峪当前水位（按Enter键结束输入）: "))
#
# current_level2 = 64.6
# target_level2 = 64.5
# # current_level2 = float(input("请输入杏石口当前水位（按Enter键结束输入）: "))


data1 = calculate_adjustment1(target_level1, current_level1)
data2 = calculate_adjustment2(target_level2, current_level2)

max_len = max(len(data1), len(data2))

# 写入数据到Excel文件
for i in range(max_len + 1):  # 0, 1, 2, 3
    # 最后一行
    if i == max_len:
        a = sheet.cell(row=start_row + i * 6, column=start_col, value=data1[len(data1)-1])
        b = sheet.cell(row=start_row + i * 6, column=start_col + 1, value=data2[len(data2)-1])
        a.font = Font(bold=True)
        b.font = Font(bold=True)
        break

    # 非最后一行
    if i >= len(data1):
        data1_now = data1[len(data1) - 1]
    else:
        data1_now = data1[i]

    if i >= len(data2):
        data2_now = data2[len(data2) - 1]
    else:
        data2_now = data2[i]

    a = sheet.cell(row=start_row + i * 6, column=start_col, value=data1_now)
    b = sheet.cell(row=start_row + i * 6, column=start_col + 1, value=data2_now)
    a.font = Font(bold=True)
    b.font = Font(bold=True)

    print("行{row:d}，列{col:d}, 数值{value:.2f}".format(row=start_row + i * 6, col=start_col, value=data1_now))
    print("行{row:d}，列{col:d}, 数值{value:.2f}".format(row=start_row + i * 6, col=start_col + 1, value=data2_now))



    # for j in range(1, 7):
    #     data_1 = data1_now - 0.005 * j
    #     data_2 = data2_now - 0.005 * j
    #     if data_1 <= target_level1:
    #         data_1 = target_level1
    #     if data_2 <= target_level2:
    #         data_2 = target_level2
    #
    #     sheet.cell(row=j + start_row + i * 6, column=start_col, value=data_1)
    #     sheet.cell(row=j + start_row + i * 6, column=start_col + 1, value=data_2)
    # for j in range(1, 7):
    #     data_1 = data1_now - 0.005 * j
    #     data_2 = data2_now - 0.005 * j
    #     if data_1 <= target_level1:
    #         data_1 = target_level1
    #     if data_2 <= target_level2:
    #         data_2 = target_level2
    #
    #     sheet.cell(row=j + start_row + i * 6, column=start_col, value=data_1)
    #     sheet.cell(row=j + start_row + i * 6, column=start_col + 1, value=data_2)
    #

    if current_level1>target_level1:
        for j in range(1, 7):
            data_1 = data1_now - 0.005 * j
            
            if data_1 <= target_level1:
                data_1 = target_level1
            
            sheet.cell(row=j + start_row + i * 6, column=start_col, value=data_1)


    if current_level2 > target_level2:
        for j in range(1, 7):
            data_2 = data2_now - 0.005 * j

            if data_2 <= target_level2:
                data_2 = target_level2

            sheet.cell(row=j + start_row + i * 6, column=start_col+1, value=data_2)



    if current_level1 < target_level1:
        for j in range(1, 7):
            data_1 = data1_now + 0.005 * j

            if data_1 >= target_level1:
                data_1 = target_level1

            sheet.cell(row=j + start_row + i * 6, column=start_col, value=data_1)



    if current_level2 < target_level2:
        for j in range(1, 7):
            data_2 = data2_now + 0.005 * j

            if data_2 >= target_level2:
                data_2 = target_level2

            sheet.cell(row=j + start_row + i * 6, column=start_col+1, value=data_2)
            
            
            
            

    # if current_level1 < target_level1:


    #
    # for j in range(1, 7):
    #     data_1 = data1_now - 0.005 * j
    #     data_2 = data2_now - 0.005 * j
    #     if data_1 <= target_level1:
    #         data_1 = target_level1
    #     if data_2 <= target_level2:
    #         data_2 = target_level2
    #
    #     sheet.cell(row=j + start_row + i * 6, column=start_col, value=data_1)
    #     sheet.cell(row=j + start_row + i * 6, column=start_col + 1, value=data_2)



# 保存Excel文件
workbook.save(file_path)

print(f"水位数据已成功写入Excel文件，从第{start_row}行第{start_col}列开始插入。")
