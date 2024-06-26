# -*- coding: UTF-8 -*-
# -*- coding: UTF-8 -*-
# -*- coding: UTF-8 -*-
import openpyxl
from openpyxl.styles import Font
import pandas as pd

# 打开现有的Excel文件  todo 修改文件名

# file_path = '调试20231018(v2).xlsx'
file_path = '../new.xlsx'
workbook = openpyxl.load_workbook(file_path)

# 选择要写入的工作表  todo 修改sheet名
# sheet = workbook['Sheet1 (2)']
sheet = workbook['6分钟分步水量控制模型']

# 让用户指定要插入数据的位置  todo 修改位置 部分位置写不进去，可以插入到文件中空白位置再粘贴
start_row = 12
start_col = 12

# 模块1和模块2的数据生成
target_level1 = 94.8
current_level1 = 94.4

target_level2 = 64.5
current_level2 = 64.4

def calculate_adjustment1(target_level1, current_level1, normal_descent=0.03):
    steps = 0
    adjusted_level = current_level1
    data1 = []

    while abs(adjusted_level - target_level1) > 0.001:
        if target_level1 > adjusted_level:
            step_adjustment = min(normal_descent, target_level1 - adjusted_level)
        else:
            step_adjustment = max(-normal_descent, target_level1 - adjusted_level)

        adjusted_level += step_adjustment
        steps += 1
        data1.append(adjusted_level)

    return data1



def calculate_adjustment2(target_level2, current_level2, normal_descent=0.03):
    steps = 0
    adjusted_level = current_level2
    data2 = []

    while abs(adjusted_level - target_level2) > 0.001:
        if target_level2 > adjusted_level:
            step_adjustment = min(normal_descent, target_level2 - adjusted_level)
        else:
            step_adjustment = max(-normal_descent, target_level2 - adjusted_level)

        adjusted_level += step_adjustment
        steps += 1
        data2.append(adjusted_level)

    return data2

data1 = calculate_adjustment1(target_level1, current_level1)
data2 = calculate_adjustment2(target_level2, current_level2)



# 写入数据到Excel文件
for i in range(len(max(data1,data2))):  # 0, 1, 2, 3
    sheet.cell(row=start_row + i * 6, column=start_col, value=data1[i])
    sheet.cell(row=start_row + i * 6, column=start_col +1, value=data2[i])

    print("行{row:d}，列{col:d}, 数值{value:.2f}".format(row=start_row + i * 6, col=start_col, value=data1[i]))
    print("行{row:d}，列{col:d}, 数值{value:.2f}".format(row=start_row + i * 6, col=start_col + 1, value=data2[i]))




# 保存Excel文件
workbook.save(file_path)

print(f"水位数据已成功写入Excel文件，从第{start_row}行第{start_col}列开始插入。")



