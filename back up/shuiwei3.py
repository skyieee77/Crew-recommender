# -*- coding: UTF-8 -*-
# 更新 calculate_adjustment1 函数
# @Project ：hww
# @File    ：unit.py
# @Author  ：Knkiss
# @Date    ：2023/10/16 19:38
# 模块1和模块2的数据生成

import openpyxl
from openpyxl.styles import Font
import pandas as pd

# 打开现有的Excel文件  todo 修改文件名

file_path = '../data/调试20231018(v2).xlsx'
# file_path = 'new.xlsx'
workbook = openpyxl.load_workbook(file_path)

# 选择要写入的工作表  todo 修改sheet名
# sheet = workbook['Sheet1 (2)']
sheet = workbook['6分钟分步水量控制模型']

# 让用户指定要插入数据的位置  todo 修改位置 部分位置写不进去，可以插入到文件中空白位置再粘贴
start_row = 12
start_col = 12

target_level1 = 94.8
current_level1 = float(input("请输入麻峪当前水位（按Enter键结束输入）: "))


# 更新 calculate_adjustment1 函数
def calculate_adjustment1(target_level1, current_level1, normal_descent=0.03, small_descent=0.005):
    steps = 0
    adjusted_level = current_level1
    data1 = []

    while abs(adjusted_level - target_level1) > 0.001:
        # 输出步长
        print(f"步长 {steps + 1}, 模块1: {adjusted_level:.2f}")

        # 输出小步长
        for i in range(5):
            adjusted_level += small_descent
            data1.append(adjusted_level)
            print(f"小步长 {i + 1}: {adjusted_level:.3f}")

        steps += 1

    return data1

target_level2 = 64.5
current_level2 = float(input("请输入杏石口当前水位（按Enter键结束输入）: "))
# 更新 calculate_adjustment2 函数

# 更新 calculate_adjustment2 函数
def calculate_adjustment2(target_level2, current_level2, normal_descent=0.03, small_descent=0.005):
    steps = 0
    adjusted_level = current_level2
    data2 = []

    while abs(adjusted_level - target_level2) > 0.001:
        # 输出步长
        print(f"步长 {steps + 1}, 模块2: {adjusted_level:.2f}")

        # 输出小步长
        for i in range(5):
            adjusted_level += small_descent
            data2.append(adjusted_level)
            print(f"小步长 {i + 1}: {adjusted_level:.3f}")

        steps += 1

    return data2


# 再次调用 calculate_adjustment 函数并保存小步长数据
data1 = calculate_adjustment1(target_level1, current_level1)
data2 = calculate_adjustment2(target_level2, current_level2)

# 写入数据到Excel文件
for i in range(max(len(data1), len(data2))):  # 0, 1, 2, 3
    if i < len(data1):
        sheet.cell(row=start_row + i * 6, column=start_col, value=data1[i])
        print(f"行{start_row + i * 6}，列{start_col}，数值{data1[i]:.2f}")

    sheet.cell(row=start_row + i * 6, column=start_col + 1, value=data2[i])
    print(f"行{start_row + i * 6}，列{start_col + 1}，数值{data2[i]:.2f}")




